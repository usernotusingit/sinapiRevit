"""Stage 1 — tidy the messy SINAPI Referência workbook into a star schema.

Reads SINAPI_Referência_2026_05.xlsx and emits parquet dimension + fact tables:
  data/dim_localidade.parquet
  data/dim_sinapi_composicao.parquet
  data/dim_sinapi_insumo.parquet
  data/fact_sinapi_custo.parquet         (composições: codigo x uf x regime)
  data/fact_sinapi_preco_insumo.parquet  (insumos:    codigo x uf x regime)

Layout facts (verified against the file):
  - Data starts at row 11 (1-based). Row 4 carries state (UF) codes, row 5 capitals.
  - Insumo sheets (ISD/ICD/ISE): 5 meta cols
        [Classificacao, Codigo, Descricao, Unidade, Origem], then 27 UF columns.
  - Composicao sheets (CSD/CCD/CSE): 4 meta cols
        [Grupo, Codigo, Descricao, Unidade], then 27 x (Custo R$, %AS) pairs.
  - In CSD/CCD/CSE the Codigo cell is a HYPERLINK(...MATCH(<code>,...)) formula;
    the integer code is regex-extracted.
Regime is taken from the sheet name: x S D = Sem Desoneracao, x C D = Com Desoneracao,
x S E = Sem Encargos.  I=Insumo, C=Composicao.
"""
import re
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
# Filenames in the zip use a non-UTF8 encoding (the "Referência" accent is mangled),
# so match by glob on the stable ASCII fragments rather than the literal name.
SRC_XLSX = next(p for p in (ROOT / "extracted").glob("SINAPI_Refer*ncia_2026_05.xlsx"))
OUT = ROOT / "data"

DATA_START_ROW = 11

INSUMO_SHEETS = {"ISD": "SD", "ICD": "CD", "ISE": "SE"}
COMPOSICAO_SHEETS = {"CSD": "SD", "CCD": "CD", "CSE": "SE"}

# Canonical SINAPI state order — identical in every sheet's column layout, regardless of
# the per-sheet banner in row 4 (the Sem-Encargos sheets omit those labels). Order taken
# from the ISD header (row 4 = UF, row 5 = capital).
UF_CODES = ["AC", "AL", "AM", "AP", "BA", "CE", "DF", "ES", "GO", "MA", "MG", "MS", "MT",
            "PA", "PB", "PE", "PI", "PR", "RJ", "RN", "RO", "RR", "RS", "SC", "SE", "SP", "TO"]
UF_CAPITALS = ["RIO BRANCO", "MACEIO", "MANAUS", "MACAPA", "SALVADOR", "FORTALEZA", "BRASILIA",
               "VITORIA", "GOIANIA", "SAO LUIS", "BELO HORIZONTE", "CAMPO GRANDE", "CUIABA",
               "BELEM", "JOAO PESSOA", "RECIFE", "TERESINA", "CURITIBA", "RIO DE JANEIRO",
               "NATAL", "PORTO VELHO", "BOA VISTA", "PORTO ALEGRE", "FLORIANOPOLIS", "ARACAJU",
               "SAO PAULO", "PALMAS"]
assert len(UF_CODES) == len(UF_CAPITALS) == 27

_CODE_RE = re.compile(r"MATCH\((\d+)")


def _extract_code(cell):
    """Return integer SINAPI code from a plain int or a HYPERLINK(...MATCH(code,...)) formula."""
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return int(cell)
    m = _CODE_RE.search(str(cell))
    if m:
        return int(m.group(1))
    s = str(cell).strip()
    return int(s) if s.isdigit() else None


def parse_insumos(wb):
    custo_rows, dim = [], {}
    for sheet, regime in INSUMO_SHEETS.items():
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            codigo = _extract_code(row[1])
            if codigo is None:
                continue
            classificacao, descricao, unidade, origem = row[0], row[2], row[3], row[4]
            dim.setdefault(codigo, (classificacao, descricao, unidade, origem))
            for k, uf in enumerate(UF_CODES):
                preco = row[5 + k]
                if preco is None or preco == "":
                    continue
                custo_rows.append((codigo, uf, regime, float(preco)))
    dim_df = pd.DataFrame(
        [(c, *v) for c, v in dim.items()],
        columns=["codigo", "classificacao", "descricao", "unidade", "origem"],
    )
    fact_df = pd.DataFrame(custo_rows, columns=["codigo", "uf", "regime", "preco"])
    return dim_df, fact_df


def parse_composicoes(wb):
    custo_rows, dim = [], {}
    for sheet, regime in COMPOSICAO_SHEETS.items():
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            codigo = _extract_code(row[1])
            if codigo is None:
                continue
            grupo, descricao, unidade = row[0], row[2], row[3]
            dim.setdefault(codigo, (grupo, descricao, unidade))
            for k, uf in enumerate(UF_CODES):
                custo = row[4 + 2 * k]
                pct_as = row[5 + 2 * k]
                if custo is None or custo == "":
                    continue
                custo_rows.append((codigo, uf, regime, float(custo),
                                   float(pct_as) if pct_as not in (None, "") else None))
    dim_df = pd.DataFrame(
        [(c, *v) for c, v in dim.items()],
        columns=["codigo", "grupo", "descricao", "unidade"],
    )
    fact_df = pd.DataFrame(custo_rows, columns=["codigo", "uf", "regime", "custo_rs", "pct_as"])
    return dim_df, fact_df


def main():
    print(f"Reading {SRC_XLSX.name} ...")
    wb = openpyxl.load_workbook(SRC_XLSX, read_only=True, data_only=False)

    dim_ins, fact_ins = parse_insumos(wb)
    dim_comp, fact_comp = parse_composicoes(wb)
    wb.close()

    dim_loc = pd.DataFrame({"uf": UF_CODES, "capital": UF_CAPITALS})

    OUT.mkdir(exist_ok=True)
    dim_loc.to_parquet(OUT / "dim_localidade.parquet", index=False)
    dim_comp.to_parquet(OUT / "dim_sinapi_composicao.parquet", index=False)
    dim_ins.to_parquet(OUT / "dim_sinapi_insumo.parquet", index=False)
    fact_comp.to_parquet(OUT / "fact_sinapi_custo.parquet", index=False)
    fact_ins.to_parquet(OUT / "fact_sinapi_preco_insumo.parquet", index=False)

    print("--- SINAPI tidy summary ---")
    print(f"dim_localidade:        {len(dim_loc):>7} states")
    print(f"dim_sinapi_composicao: {len(dim_comp):>7} composicoes")
    print(f"dim_sinapi_insumo:     {len(dim_ins):>7} insumos")
    print(f"fact_sinapi_custo:     {len(fact_comp):>7} rows (codigo x uf x regime)")
    print(f"fact_preco_insumo:     {len(fact_ins):>7} rows")
    print(f"regimes (comp): {sorted(fact_comp.regime.unique())}")
    print(f"MG composicoes priced (SD): "
          f"{len(fact_comp[(fact_comp.uf=='MG') & (fact_comp.regime=='SD')])}")


if __name__ == "__main__":
    main()
